#!/usr/bin/env python3

"""
usage: xlslisp.py [-h] [-f] FILE

export sheets of traces of excel lisp into git change tracking, via pandas

positional arguments:
  FILE         the xlsx file to read

optional arguments:
  -h, --help   show this help message and exit
  -f, --force  ask less questions

dependencies:

  python3 -m venv --prompt PIPS pips
  source pips/bin/activate
  pip install --upgrade wheel
  pip install --upgrade pip

  pip install --upgrade pandas==1.2.5  # last stable of Pandas PyData Org
  pip install --upgrade openpyxcl

examples:

  xlslisp.py -f like-py.xlsx && ls like-py-*.csv
"""

# code reviewed by people, and by Black and Flake8 bots


import __main__
import argparse
import csv
import datetime as dt
import difflib
import io
import os
import pdb
import string
import sys


import openpyxl  # called by "pd.read_excel" for workbooks of multiple sheets


_ = pdb


def main():
    """Run from the command line"""

    era = dt.datetime.now()

    parser = xlslisp_compile_argdoc()
    args = parser.parse_args()

    space = os.path.splitext(args.file)[0]

    # Import the Values of Sheets of one Xlsx File

    sheet_by_name = openpyxl.load_workbook(args.file, data_only=True)
    sheet_by_name_keys_list = sheet_by_name.sheetnames

    stderr_print(
        "xlslisp: reading {} sheets from:  {}".format(
            len(sheet_by_name_keys_list), args.file
        )
    )

    # Option to quit early

    if not args.force:
        stderr_print(
            "xlslisp.py:  Xlsx imported, run again with --force to replace Csv's"
        )

        sys.exit(1)

    # Visit each Sheet

    for (index, sheetname) in enumerate(sheet_by_name_keys_list):
        sheet = sheet_by_name[sheetname]

        csv_name = "{space}-{dashed_sheet}.csv".format(
            space=space, dashed_sheet=sheetname.replace(" ", "-")
        ).lower()

        # Collect Rows of String Values

        csv_ragged_rows = list()
        for row_index in range(sheet.max_row):
            row_mark = 1 + row_index

            csv_cells = list()

            for col_index in range(sheet.max_column):
                cell = sheet.cell(1 + row_index, 1 + col_index)
                col_mark = cell.column_letter
                assert col_mark == excel_az_mark(col_index)

                if False:
                    if (col_mark, row_mark) == ("C", 89):
                        pdb.set_trace()

                csv_cells.append(cell.value)

                # Warn of trailing spaces

                if str(csv_cells[-1]).endswith(" "):
                    stderr_print(
                        "xlslisp: Warning: "
                        "could rstrip cell at: {!r}!{}{} {}".format(
                            sheetname, col_mark, row_mark, csv_cells[-1]
                        )
                    )

            csv_ragged_rows.append(csv_cells)

        # Format as rectangular Csv to please GitHub
        #
        # per GitHub > Rendering CSV and TSV data
        # flagging ragged as "we can make this file beautiful and searchable"
        #

        csv_rows = rows_complete(csv_ragged_rows, cell=None)

        charstream = io.StringIO()
        csv_writer = csv.writer(charstream)
        for csv_cells in csv_rows:
            csv_writer.writerow(csv_cells)

        charstream.seek(0)
        csv_chars = charstream.read()

        # Write the lines with local "os.linesep" line-ending's
        # specifically Not the mix of "\r\n" and "\n" from multi-line Excel cells
        # but without rstrip'ping the lines  # TODO: poor choice to skip rstrip?

        csv_lines = csv_chars.splitlines()
        csv_joined = "\n".join(csv_lines) + "\n"

        stderr_print(
            "xlslisp: writing {} chars of {} rows to:  {}".format(
                len(csv_joined), sheet.max_row, csv_name
            )
        )

        with open(csv_name, "w") as csv_writing:
            csv_writing.write(csv_joined)

    now = dt.datetime.now()
    stderr_print("xlslisp: elapsed time of", (now - era), "since", era)

    sys.exit(0)


def xlslisp_compile_argdoc():
    """Parse the command-line args per top-of-file arg doc"""

    doc = __main__.__doc__

    prog = doc.strip().splitlines()[0].split()[1]
    description = list(_ for _ in doc.strip().splitlines() if _)[1]
    epilog_at = doc.index("dependencies:")
    epilog = doc[epilog_at:]

    parser = argparse.ArgumentParser(
        prog=prog,
        description=description,
        add_help=True,
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=epilog,
    )

    parser.add_argument("file", metavar="FILE", help="the xlsx file to read")
    parser.add_argument("-f", "--force", action="count", help="ask less questions")

    exit_unless_main_doc_eq(parser)

    return parser


# deffed in many files  # missing from docs.python.org
def excel_az_mark(row_index):
    """Map row indices 0 1 2 ... to Excel A .. Z, AA .. ZZ, AAA .. XFD ..."""

    marks = ""

    az = string.ascii_uppercase
    base = len(az)

    width = 1  # width 1 for Excel A .. Z, 2 for AA .. ZZ, etc
    floor = 0  # min index of this width
    ceil = base  # one beyond the max index of this width
    while ceil <= row_index:
        width += 1
        floor = ceil
        ceil = floor + (base ** width)

    remainder = row_index - floor
    for power in range(width):
        (remainder, digit) = divmod(remainder, base)
        marks = az[digit] + marks

    return marks  # "XFD" at 16383 aka ((1 << 14) - 1)


# deffed in many files  # missing from docs.python.org
def exit_unless_main_doc_eq(parser):
    """Exit nonzero now, unless __main__.__doc__ equals "parser.format_help()" """

    file_filename = os.path.split(__file__)[-1]

    main_doc = __main__.__doc__.strip()
    parser_doc = parser.format_help()

    got = main_doc
    got_filename = "./{} --help".format(file_filename)
    want = parser_doc
    want_filename = "argparse.ArgumentParser(..."

    diff_lines = list(
        difflib.unified_diff(
            a=got.splitlines(),
            b=want.splitlines(),
            fromfile=got_filename,
            tofile=want_filename,
        )
    )

    if diff_lines:
        lines = list((_.rstrip() if _.endswith("\n") else _) for _ in diff_lines)
        stderr_print("\n".join(lines))

        sys.exit(1)


# deffed in many files  # missing from docs.python.org
def rows_complete(rows, cell):
    """Add cells to make every row as wide as the widest row"""

    completed_rows = list()
    if rows:
        max_row_width = max(len(_) for _ in rows)

        for row in rows:
            completed_row = row + ((max_row_width - len(row)) * [cell])
            completed_rows.append(completed_row)

    return completed_rows


# deffed in many files  # missing from docs.python.org
def stderr_print(*args, **kwargs):
    """Like Print, but flush don't write Stdout and do write and flush Stderr"""

    sys.stdout.flush()
    print(*args, **kwargs, file=sys.stderr)
    sys.stderr.flush()

    # else caller has to "{}\n".format(...) and flush


if __name__ == "__main__":
    main()


# copied from:  git clone https://github.com/pelavarre/like-py-xlsx.git
